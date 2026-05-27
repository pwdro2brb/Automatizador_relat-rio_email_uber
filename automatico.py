import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import unicodedata
import re
import os
 
# =========================
# FORMATAÇÃO BRL (R$) - COLUNA POR CABEÇALHO + COLUNA K
# =========================
def to_float(v):
    """Converte valores do Excel/pandas para número, aceitando '1.234,56' e '1234.56'."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()

    # remove "R$" e espaços
    s = s.replace("R$", "").strip()

    # se vier no padrão BR: 1.234,56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    # se vier só com vírgula: 1234,56
    elif "," in s:
        s = s.replace(",", ".")

    try:
        return float(s)
    except:
        return None


def sanitize_filename(name: str) -> str:
    """Remove caracteres inválidos para nome de arquivo no Windows."""
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]+', '-', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name[:150]  # evita caminho longo

def sanitize_sheetname(name: str) -> str:
    """Nome de aba no Excel: máx 31 chars e sem : \\ / ? * [ ]"""
    name = str(name).strip()
    name = re.sub(r'[:\\/?*\\[\\]]+', '-', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name[:31] if name else "Planilha1"

def find_col_by_prefix(df_columns, wanted: str):
    """
    Acha coluna por match exato ou por prefixo.
    (Resolve casos tipo 'Endereço de destino+P:...').
    """
    w = norm(wanted)
    for c in df_columns:
        if norm(c) == w:
            return c
    for c in df_columns:
        if norm(c).startswith(w):
            return c
    return None

def criar_planilhas_por_responsavel(
    consolidado_path: str,
    teste_macro_path: str,
    pasta_mes: str = None
):
    # Pasta do mês atual: "2026,04" (mesmo padrão dos seus arquivos)
    now = datetime.now()
    if pasta_mes is None:
        pasta_mes = f"{now.year},{now.month:02d}"

    os.makedirs(pasta_mes, exist_ok=True)

    # Lê TESTE MACRO (coluna A = Responsavel)
    macro_df = pd.read_excel(teste_macro_path, engine="openpyxl")
    macro_df.columns = [str(c).strip() for c in macro_df.columns]
    col_resp_macro = macro_df.columns[0]  # "Responsavel"
    responsaveis = (
        macro_df[col_resp_macro]
        .dropna()
        .astype(str)
        .map(lambda x: x.strip())
        .loc[lambda s: s != ""]
        .unique()
        .tolist()
    )

    # Lê consolidado
    cons_df = pd.read_excel(consolidado_path, engine="openpyxl")
    cons_df.columns = [str(c).strip() for c in cons_df.columns]

    # Colunas que você pediu
    col_end_dest = find_col_by_prefix(cons_df.columns, "Endereço de destino")
    col_nome_aj = find_col_by_prefix(cons_df.columns, "COLABORADOR AJUSTADO")
    col_resp_cc = find_col_by_prefix(cons_df.columns, "RESPONSÁVEL CC")

    # Lista final de colunas na ordem correta
    colunas_saida = [
        "Data da solicitação (local)",
        "Hora da solicitação (local)",
        "Data de chegada (local)",
        "Hora de chegada (local)",
        "Nome",
        "Sobrenome",
        "Nome ajustado",  # vira COLABORADOR AJUSTADO
        "E-mail",
        "Cidade",
        "Distância (mi)",
        "Duração (min)",
        "Endereço de partida",
        "Endereço de destino",
        "Código da despesa",
        "Responsavel",  # vira RESPONSÁVEL CC
        "Detalhamento da despesa",
        "Valor da transação em BRL (com tributos)",
    ]

    # Estilos (baseado nas suas imagens)
    roxo = PatternFill("solid", fgColor="7030A0")     # roxo (Excel padrão)
    laranja = PatternFill("solid", fgColor="ED7D31")  # laranja (Excel padrão)
    font_white_bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for resp in responsaveis:
        # Filtra as linhas daquele responsável (match normalizado)
        resp_norm = norm(resp)
        if col_resp_cc is None:
            # se por algum motivo não existir a coluna
            continue

        f = cons_df[col_resp_cc].fillna("").astype(str).map(norm) == resp_norm
        df_r = cons_df.loc[f].copy()

        # Se não houver uso para aquele responsável, pula (extra segurança)
        if df_r.empty:
            continue

        # Monta dataframe com as colunas desejadas
        saida = pd.DataFrame()

        # colunas diretas
        for c in [
            "Data da solicitação (local)",
            "Hora da solicitação (local)",
            "Data de chegada (local)",
            "Hora de chegada (local)",
            "Nome",
            "Sobrenome",
            "E-mail",
            "Cidade",
            "Distância (mi)",
            "Duração (min)",
            "Endereço de partida",
            "Código da despesa",
            "Detalhamento da despesa",
            "Valor da transação em BRL (com tributos)",
        ]:
            if c in df_r.columns:
                saida[c] = df_r[c]
            else:
                saida[c] = ""

        # Endereço de destino (pode estar com sufixo estranho)
        saida["Endereço de destino"] = df_r[col_end_dest] if col_end_dest else ""

        # Nome ajustado = COLABORADOR AJUSTADO
        saida["Nome ajustado"] = df_r[col_nome_aj] if col_nome_aj else ""

        # Responsavel = RESPONSÁVEL CC
        saida["Responsavel"] = df_r[col_resp_cc] if col_resp_cc else ""

        # Reordena exatamente como você definiu
        saida = saida[colunas_saida]

        # Cria workbook do responsável
        wb = Workbook()
        ws = wb.active
        ws.title = sanitize_sheetname(resp)

        # Escreve dados
        for row in dataframe_to_rows(saida, index=False, header=True):
            ws.append(row)

        # Congela topo
        ws.freeze_panes = "A2"

        # Ativa filtro na linha 1
        ws.auto_filter.ref = ws.dimensions

        # Aplica estilo no cabeçalho:
        # - Roxo em tudo
        # - Laranja apenas em "Nome ajustado" e "Responsavel"
        header_row = 1
        header_map = {ws.cell(header_row, i).value: i for i in range(1, ws.max_column + 1)}

        for col_name, idx in header_map.items():
            cell = ws.cell(header_row, idx)
            cell.fill = roxo
            cell.font = font_white_bold
            cell.alignment = center
            cell.border = border

        for special in ["Nome ajustado", "Responsavel"]:
            if special in header_map:
                cell = ws.cell(header_row, header_map[special])
                cell.fill = laranja
                cell.font = font_white_bold
                cell.alignment = center
                cell.border = border
        
        # =========================
        # BORDA EM TODAS AS CÉLULAS COM CONTEÚDO
        # =========================
        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                if cell.value not in (None, ""):
                    cell.border = border

        # =========================
        # ALINHAMENTO DO CORPO (A PARTIR DA LINHA 2)
        # Colunas:
        # A–E, I–K, N, Q
        # =========================
        # Conjuntos de colunas (índices 1-based)
        cols_centralizar = (
            list(range(1, 6)) +      # A–E
            list(range(9, 12)) +     # I–K
            [14] +                   # N
            [17]                     # Q
        )

        for r in range(2, max_row + 1):  # a partir da segunda linha
            for c in cols_centralizar:
                if c <= max_col:
                    cell = ws.cell(r, c)
                    if cell.value not in (None, ""):
                        cell.alignment = center

        # Ajuste opcional de altura do cabeçalho
        ws.row_dimensions[1].height = 50

        # Modificar altura das linhas do corpo para 15 (opcional, ajuda a ficar parecido com o modelo)
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 15

        
        BRL_FORMAT = '"R$" #,##0.00'

        # 1) Formatar a coluna que tem o cabeçalho do valor em BRL (recomendado)
        header_map = {ws.cell(1, i).value: i for i in range(1, ws.max_column + 1)}
        col_valor_brl = header_map.get("Valor da transação em BRL (com tributos)")

        if col_valor_brl:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, col_valor_brl)
                val = to_float(cell.value)
                if val is not None:
                    cell.value = val
                    cell.number_format = BRL_FORMAT

        col_k = 17
        if col_k <= ws.max_column:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, col_k)
                val = to_float(cell.value)
                if val is not None:
                    cell.value = val
                    cell.number_format = BRL_FORMAT


        # Salva arquivo na pasta do mês
        file_name = sanitize_filename(resp) + ".xlsx"
        full_path = os.path.join(pasta_mes, file_name)
        wb.save(full_path)

    print(f"✅ Pasta criada: {pasta_mes}")
    print("✅ Arquivos por responsável gerados com sucesso.")


uber_path = "Consolidado para envio.xlsx"
resp_path = "Responsaveis Por Centro de Custos.xlsx"
ativos_path = "Base de Ativos 14.05.2026.xlsx"

CARGO_NAO_ENCONTRADO = "cargo não encontrado"
MSG_CC_MISSING = "CC NÃO ESTÁ MAIS NA BASE UBER"

BANIDOS_NOMES = {
    "JUNIA GALVAO",
    "THIAGO CORREA ELY",
    "RAFAEL PIRES E ALBUQUERQUE",
    "RONALDO PEDREIRA AYRES DA MOTTA FILHO",
    "RICARDO PAIXAO PINTO RODRIGUES",
    "RAPHAEL ROCHA LAFETA",
    "EDUARDO FISCHER TEIXEIRA DE SOUZA",
    "RODRIGO MARTINS DE RESENDE",
    "RAFAEL NAZARETH MENIN TEIXEIRA DE SOUZA",
    "RUBENS MENIN TEIXEIRA DE SOUZA"
}
CARGOS_PROIBIDOS_CONTEM = ["DIRETOR EXECUTIVO", "PRESIDENTE", "CONSELHEIRO"]

def norm(s):
    if s is None:
        return ""
    if isinstance(s, float) and pd.isna(s):
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split())

def norm_email(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"^mailto:", "", s)
    s = s.replace("[", "").replace("]", "").strip()
    return s

def is_excluded(name, cargo):
    n = norm(name)
    c = norm(cargo)
    if n in {norm(x) for x in BANIDOS_NOMES}:
        return True
    for sub in [norm(x) for x in CARGOS_PROIBIDOS_CONTEM]:
        if sub in c:
            return True
    return False

# ---------------------------
# FUNÇÕES PARA NÃO ALTERAR DATA/HORA (transformar em TEXTO)
# ---------------------------
def fmt_date_as_text(x):
    """Converte datas para mm/dd/yyyy como TEXTO."""
    if pd.isna(x):
        return ""
    dt = pd.to_datetime(x, errors="coerce")
    if pd.isna(dt):
        return str(x).strip()
    return dt.strftime("%d/%m/%Y")

def fmt_time_as_text(x):
    """Converte horas para h:mmAM/PM como TEXTO (sem zero à esquerda)."""
    if pd.isna(x):
        return ""
    dt = pd.to_datetime(x, errors="coerce")
    if pd.isna(dt):
        s = str(x).strip()
        return s
    # exemplo: 08:22AM -> 8:22AM
    out = dt.strftime("%I:%M%p")
    return out.lstrip("0")

# =========================
# 1) RESPONSÁVEIS
# =========================
resp_df = pd.read_excel(resp_path, engine="openpyxl")
resp_df.columns = [str(c).strip() for c in resp_df.columns]

col_cc = next((c for c in resp_df.columns if norm(c) in {
    "CODIGODEDESPESAS", "CODIGO DE DESPESAS", "CÓDIGO DE DESPESAS",
    "CODIGO DA DESPESA", "CÓDIGO DA DESPESA"
}), resp_df.columns[0])

col_resp = next((c for c in resp_df.columns if norm(c) in {"RESPONSAVEL", "RESPONSÁVEL"}), None)
if col_resp is None:
    col_resp = next((c for c in resp_df.columns if "RESP" in norm(c)), None)

col_email = next((c for c in resp_df.columns if "MAIL" in norm(c)), None)
col_cargo_resp = next((c for c in resp_df.columns if "CARGO" in norm(c)), None)
if col_cargo_resp is None and len(resp_df.columns) >= 4:
    col_cargo_resp = resp_df.columns[3]

resp_df = resp_df.dropna(subset=[col_cc]).copy()
resp_df["_cc_norm"] = resp_df[col_cc].astype(str).map(norm)
resp_df["_resp_norm"] = resp_df[col_resp].astype(str).map(norm)
resp_df["_email_norm"] = resp_df[col_email].map(norm_email)

cc_to_resp = dict(zip(resp_df["_cc_norm"], resp_df[col_resp].astype(str)))
cc_to_email = dict(zip(resp_df["_cc_norm"], resp_df[col_email].astype(str)))
resp_name_to_cargo = resp_df.groupby("_resp_norm")[col_cargo_resp].first().to_dict()
resp_name_to_email_norm = resp_df.groupby("_resp_norm")["_email_norm"].first().to_dict()

# =========================
# 2) BASE ATIVOS (email->cargo / nome->cargo)
# =========================
wb = load_workbook(ativos_path, read_only=True, data_only=True)
ws = wb.worksheets[0]

header_row_idx = None
for r in range(1, 60):
    row = [ws.cell(r, c).value for c in range(1, 40)]
    if "NOME FUNCIONARIO" in [norm(v) for v in row] or "NOME FUNCIONÁRIO" in [norm(v) for v in row]:
        header_row_idx = r
        header_vals = row
        break

if header_row_idx is None:
    header_row_idx = 7
    header_vals = [ws.cell(header_row_idx, c).value for c in range(1, 40)]

header_norm = [norm(v) for v in header_vals]

def find_col(possible):
    for p in possible:
        p = norm(p)
        if p in header_norm:
            return header_norm.index(p) + 1
    return None

col_nome_func = find_col(["Nome Funcionário", "Nome Funcionario", "Nome"]) or 2
col_email_func = find_col(["E-mail", "Email", "E Mail"]) or 5
col_cargo_func = find_col(["Nome da Função", "Nome da Funcao", "Função", "Funcao", "Cargo"]) or 7

email_to_cargo = {}
name_to_cargo = {}

for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
    nome = row[col_nome_func - 1] if col_nome_func - 1 < len(row) else None
    email = row[col_email_func - 1] if col_email_func - 1 < len(row) else None
    cargo = row[col_cargo_func - 1] if col_cargo_func - 1 < len(row) else None
    cargo_str = "" if cargo is None else str(cargo)

    if email:
        e = norm_email(email)
        if e and (e not in email_to_cargo or (not email_to_cargo[e] and cargo_str)):
            email_to_cargo[e] = cargo_str

    if nome:
        n = norm(nome)
        if n and (n not in name_to_cargo or (not name_to_cargo[n] and cargo_str)):
            name_to_cargo[n] = cargo_str

wb.close()

# =========================
# 3) UBER (consolidado)
# =========================
uber_df = pd.read_excel(uber_path, engine="openpyxl")
uber_df.columns = [str(c).strip() for c in uber_df.columns]

# remove Unnamed
unnamed_cols = [c for c in uber_df.columns if str(c).startswith("Unnamed")]
if unnamed_cols:
    uber_df = uber_df.drop(columns=unnamed_cols)

# remove coluna UTC completa se existir
drop_col = next((c for c in uber_df.columns if norm(c) == norm("Registro de data e hora da transação (UTC)")), None)
if drop_col:
    uber_df = uber_df.drop(columns=[drop_col])

col_cc_uber = next((c for c in uber_df.columns if norm(c) in {
    "CODIGO DA DESPESA", "CÓDIGO DA DESPESA", "CODIGO DE DESPESA", "CÓDIGO DE DESPESA"
}), None)
col_email_uber = next((c for c in uber_df.columns if "MAIL" in norm(c)), None)
col_nome = next((c for c in uber_df.columns if norm(c) == "NOME"), None)
col_sobrenome = next((c for c in uber_df.columns if norm(c) == "SOBRENOME"), None)

if col_cc_uber is None:
    raise ValueError("Não encontrei a coluna 'Código da despesa' no Uber.")
if col_email_uber is None:
    raise ValueError("Não encontrei a coluna de e-mail no Uber.")
if col_nome is None or col_sobrenome is None:
    raise ValueError("Não encontrei 'Nome' e/ou 'Sobrenome' no Uber.")

# --- PRESERVAÇÃO DE DATA/HORA COMO TEXTO ---
# Converte TODAS as colunas que começam com Data/Hora para texto formatado
for c in uber_df.columns:
    if norm(c).startswith("DATA "):
        uber_df[c] = uber_df[c].map(fmt_date_as_text)
    if norm(c).startswith("HORA "):
        uber_df[c] = uber_df[c].map(fmt_time_as_text)

uber_df["_cc_norm"] = uber_df[col_cc_uber].astype(str).map(norm)

uber_df["RESPONSÁVEL CC"] = uber_df["_cc_norm"].map(lambda x: cc_to_resp.get(x, MSG_CC_MISSING))

def cargo_responsavel(resp_name, cc_norm):
    rn = norm(resp_name)
    if resp_name == MSG_CC_MISSING or rn == norm(MSG_CC_MISSING):
        return ""
    c = resp_name_to_cargo.get(rn, "")
    if c:
        return c
    email_resp = cc_to_email.get(cc_norm, "")
    c2 = email_to_cargo.get(norm_email(email_resp), "")
    if c2:
        return c2
    return name_to_cargo.get(rn, "")

uber_df["CARGO RESPONSÁVEL CC"] = [
    cargo_responsavel(r, cc) for r, cc in zip(uber_df["RESPONSÁVEL CC"], uber_df["_cc_norm"])
]

uber_df["COLABORADOR AJUSTADO"] = (
    uber_df[col_nome].fillna("").astype(str).str.strip() + " " +
    uber_df[col_sobrenome].fillna("").astype(str).str.strip()
).str.replace(r"\s+", " ", regex=True).str.strip()

def cargo_colaborador(email, nome):
    c = email_to_cargo.get(norm_email(email), "")
    if c:
        return c
    c2 = name_to_cargo.get(norm(nome), "")
    if c2:
        return c2
    return CARGO_NAO_ENCONTRADO

uber_df["CARGO COLABORADOR"] = [
    cargo_colaborador(e, n) for e, n in zip(uber_df[col_email_uber], uber_df["COLABORADOR AJUSTADO"])
]

uber_df["OBSERVAÇÃO"] = ""

anchor = next((c for c in uber_df.columns if norm(c) == norm("Valor da transação em BRL (com tributos)")), None)
new_cols = ["RESPONSÁVEL CC", "CARGO RESPONSÁVEL CC", "COLABORADOR AJUSTADO", "CARGO COLABORADOR", "OBSERVAÇÃO"]
cols = [c for c in uber_df.columns if c not in new_cols]

if anchor and anchor in cols:
    idx = cols.index(anchor) + 1
    cols = cols[:idx] + new_cols + cols[idx:]
else:
    cols += new_cols

uber_out_df = uber_df[cols].drop(columns=["_cc_norm"], errors="ignore")

# =========================
# 4) SALVA CONSOLIDADO + APLICA ESTILOS
# =========================
out_consolidado = "consolidado_para_envio_ATUALIZADO_COM_ESTILO.xlsx"

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "consolidado para envio"

for row in dataframe_to_rows(uber_out_df, index=False, header=True):
    ws_out.append(row)

ws_out.freeze_panes = "A2"

# --- AUTO FILTRO (para aparecer o dropdown do filtro) ---
ws_out.auto_filter.ref = ws_out.dimensions

# --- ESTILOS (headers) ---
# Cores aproximadas conforme as imagens

fill_azul   = PatternFill("solid", fgColor="B7DEE8")   # azul claro
fill_laranja= PatternFill("solid", fgColor="FFC000")   # laranja
fill_rosa   = PatternFill("solid", fgColor="FCE4D6")   # rosa/pêssego claro

bold_font = Font(bold=True, color="000000")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="A6A6A6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Mapeia cabeçalho -> índice (coluna)
header = [cell.value for cell in ws_out[1]]
col_index = {str(h).strip(): i + 1 for i, h in enumerate(header) if h is not None}

def find_col_idx(col_name: str):
    """
    Encontra o índice da coluna:
    1) match exato (normalizado)
    2) match por prefixo (normalizado) - útil p/ 'Endereço de destino+P:...'
    """
    target = norm(col_name)
    for h, idx in col_index.items():
        if norm(h) == target:
            return idx
    for h, idx in col_index.items():
        if norm(h).startswith(target):
            return idx
    return None

def style_header(col_name, fill):
    idx = find_col_idx(col_name)
    if idx is None:
        return
    c = ws_out.cell(row=1, column=idx)
    c.fill = fill
    c.font = bold_font
    c.alignment = center
    c.border = border

# 1) Colunas que você quer no "azul" (inclui Valor extra em BRL)
AZUL_HEADERS = [
    "Data da solicitação (UTC)",
    "Hora da solicitação (UTC)",
    "Data da solicitação (local)",
    "Hora da solicitação (local)",
    "Data de chegada (UTC)",
    "Hora de chegada (UTC)",
    "Data de chegada (local)",
    "Hora de chegada (local)",
    "Compensação do fuso horário de solicitação a partir do UTC",
    "Nome",
    "Sobrenome",
    "E-mail",
    "ID do funcionário",
    "Serviço",
    "Cidade",
    "Distância (mi)",
    "Duração (min)",
    "Endereço de partida",
    "Endereço de destino",  # pega mesmo se estiver 'Endereço de destino+P:...'
    "Detalhamento da despesa",
    "Faturas",
    "Programa",
    "Grupo",
    "Forma de pagamento",
    "Valor na moeda local (sem tributos)",
    "Tributos na moeda local",
    "Valor extra em moeda local",
    "Valor da transação na moeda local (com tributos)",
    "Código da moeda local",
    "Valor em BRL (sem tributos)",
    "Tributos em BRL",
    "Valor extra em BRL",
]

for col_name in AZUL_HEADERS:
    style_header(col_name, fill_azul)

# 2) Rosa/pêssego (essas duas você pediu iguais à 3ª imagem)
style_header("Código da despesa", fill_rosa)
style_header("Tipo de transação", fill_rosa)
style_header("Valor da transação em BRL (com tributos)", fill_rosa)

# 3) Laranja (colunas criadas)
for col_name in [
    "RESPONSÁVEL CC",
    "CARGO RESPONSÁVEL CC",
    "COLABORADOR AJUSTADO",
    "CARGO COLABORADOR",
    "OBSERVAÇÃO",
]:
    style_header(col_name, fill_laranja)

# Altura do header (opcional, ajuda a ficar parecido com o seu modelo)
ws_out.row_dimensions[1].height = 22


wb_out.save(out_consolidado)

print("OK! Arquivo gerado:", out_consolidado)

# =========================
# 5) TESTE MACRO + ENVIAR EMAIL (somente CC usado)
# =========================
used_resp = uber_out_df["RESPONSÁVEL CC"].fillna("").astype(str).str.strip()
used_resp = used_resp[(used_resp != "") & (used_resp != MSG_CC_MISSING)]
used_norm = set(used_resp.map(norm))

macro_rows = []
for rn in sorted(used_norm):
    orig = resp_df.loc[resp_df["_resp_norm"] == rn, col_resp].dropna()
    orig_name = orig.iloc[0] if len(orig) > 0 else rn.title()

    email = resp_name_to_email_norm.get(rn, "")
    cargo = resp_name_to_cargo.get(rn, "")
    if not cargo and email:
        cargo = email_to_cargo.get(email, "")
    if not cargo:
        cargo = name_to_cargo.get(rn, "")

    if is_excluded(orig_name, cargo):
        continue
    if not email:
        continue

    macro_rows.append({"Responsavel": orig_name, "E-mail": email})

macro_unique = pd.DataFrame(macro_rows)
if not macro_unique.empty:
    macro_unique["_email_norm"] = macro_unique["E-mail"].map(norm_email)
    macro_unique = macro_unique.drop_duplicates("_email_norm", keep="first").drop(columns=["_email_norm"])

now = datetime.now()
prev_month = now.month - 1
prev_year = now.year
if prev_month == 0:
    prev_month = 12
    prev_year -= 1

macro_filename = f"TESTE MACRO {prev_year},{prev_month:02d} COM_ESTILO.xlsx"
email_filename = "Enviar_e-mail original COM_ESTILO.xlsx"

orange = PatternFill("solid", fgColor="FFA500")
header_font = Font(bold=True, color="FFFFFF")

# TESTE MACRO
wb_m = Workbook()
ws_m = wb_m.active
ws_m.title = "TESTE MACRO"
ws_m["A1"] = "Responsavel"
ws_m["B1"] = "E-mail"
for cell in ("A1", "B1"):
    ws_m[cell].fill = orange
    ws_m[cell].font = header_font

for i, r in enumerate(macro_unique.itertuples(index=False), start=2):
    ws_m.cell(i, 1).value = r[0]
    ws_m.cell(i, 2).value = r[1]

wb_m.save(macro_filename)

# ENVIAR EMAIL
wb_e = Workbook()
ws_e = wb_e.active
ws_e.title = "Enviar_e-mail original"
ws_e["A1"] = "E-mail"
ws_e["B1"] = "Responsavel"
ws_e["C1"] = "Corpo"
for cell in ("A1", "B1", "C1"):
    ws_e[cell].fill = orange
    ws_e[cell].font = header_font

body_text = (
    "Segue em anexo as utilizações do Uber coorporativo referente aos meses Janeiro e Fevereiro de 2026, "
    "solicitados nos centros de custos sob sua responsabilidade. Caso estiver de acordo, não é necessário responder esse e-mail."
)

for i, r in enumerate(macro_unique.itertuples(index=False), start=2):
    ws_e.cell(i, 1).value = r[1]  # email
    ws_e.cell(i, 2).value = r[0]  # nome
    ws_e.cell(i, 3).value = body_text

wb_e.save(email_filename)

print("OK! Arquivos gerados:")
print("-", macro_filename)
print("-", email_filename)


criar_planilhas_por_responsavel(
    consolidado_path="Consolidado para envio.xlsx",
    teste_macro_path=macro_filename  # ou o nome fixo do seu arquivo TESTE MACRO
)



# =========================
# GERAR PENDENCIAS_CARGO.xlsx
# =========================
PEND_FILE = "PENDENCIAS_CARGO.xlsx"

# garante que as colunas existam (se alguma não existir, cria vazia)
cols_needed = [
    "COLABORADOR AJUSTADO",
    "E-mail",
    "RESPONSÁVEL CC",
    "Código da despesa",
    "Cidade",
    "Serviço",
    "CARGO COLABORADOR",
    "Detalhamento da despesa",
    "Valor da transação em BRL (com tributos)",
]
for c in cols_needed:
    if c not in uber_out_df.columns:
        uber_out_df[c] = ""

pend_df = uber_out_df.loc[
    uber_out_df["CARGO COLABORADOR"].fillna("").astype(str).str.strip().str.lower() == "cargo não encontrado"
].copy()

# se quiser só as principais (enxuto)
pend_out_cols = [
    "COLABORADOR AJUSTADO",
    "E-mail",
    "Cidade",
    "Serviço",
    "Código da despesa",
    "RESPONSÁVEL CC",
    "CARGO COLABORADOR",
]
pend_df = pend_df[pend_out_cols]

# opcional: contagem de ocorrências por pessoa (ajuda muito)
resumo = (
    pend_df.groupby(["COLABORADOR AJUSTADO", "E-mail"], dropna=False)
    .size()
    .reset_index(name="Qtde ocorrências")
    .sort_values("Qtde ocorrências", ascending=False)
)

with pd.ExcelWriter(PEND_FILE, engine="openpyxl") as writer:
    pend_df.to_excel(writer, index=False, sheet_name="Pendencias (linhas)")
    resumo.to_excel(writer, index=False, sheet_name="Resumo por pessoa")

print(f"✅ Arquivo gerado: {PEND_FILE}  | Linhas: {len(pend_df)} | Pessoas: {len(resumo)}")